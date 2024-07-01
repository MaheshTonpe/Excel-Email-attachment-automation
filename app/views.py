from django.shortcuts import render
from django.http.response import Http404
import pandas as pd
from io import BytesIO
from app.tasks import email_background_worker
from django.core.mail import EmailMessage
from django.conf import settings
from openpyxl.styles import Font
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from datetime import datetime, timedelta
import calendar



def home(request):
    """
    column-list:
    Month Email GSTIN/UIN Party_Name Docu_Type Inv no/CreditNote/Debit No Date Rate TAXABLE_VALUE IGST_Amount SGST_Amount Total_Tax Place_of_Success Remark
    """
    # step 1; column clean -> remove white space and trim and reassign
    # step 2; read email and invoice row. 
    # step 3; convert row to excel and attach to email 
    # step 4; send email via background tasks
    
    context = {}
    email_tasks = []

    if request.method == "POST":
        try:
            files = request.FILES

            excel = files.get('file')
            if not excel:
                raise Http404('Invalid excel file')
            
            df = pd.read_excel(excel)
            # step 1
            df.columns = df.columns.str.lower().str.strip().str.replace(' ', '_')
            df.columns = df.columns.str.upper().str.replace('_', ' ')

            required_columns = [
                'MONTH', 'EMAIL', 'GSTIN/UIN', 'PARTY NAME', 'DOCU TYPE',
                'INV NO/CREDITNOTE/DEBIT NO', 'DATE', 'RATE', 'TAXABLE VALUE',
                'IGST AMOUNT','CGST AMOUNT', 'SGST AMOUNT', 'TOTAL TAX', 'PLACE OF SUPPLY', 'REMARK', 'CC MEMBERS'
            ]

            # missing_columns = [col for col in required_columns if col not in df.columns]
            # if missing_columns:
            #     raise ValueError(f"Missing columns in the uploaded file: {', '.join(missing_columns)}")


            # #To Calculate the current fiscal year range
            # today = datetime.today()
            # if today.month >= 4:  # Fiscal year starts in April
            #     start_year = today.year
            #     end_year = today.year + 1
            # else:
            #     start_year = today.year - 1
            #     end_year = today.year
            
            # fiscal_year_range = f"{start_year}-{str(end_year)[-2:]}"
            # fiscal_year = f"F.Y. {start_year}-{end_year}"
            
            # step 2
            for index, row in df.iterrows():
                heading = [
                        "SHREE GANESH PRESS N COAT INDUSTRIES PRIVATE LIMITED.\n",
                        f"{row['GSTIN/UIN']}.\n"
                        "Report Name: Not in GSTR 2B.\n"
                        "April 23 to April 24.\n\n"
                        # f"{fiscal_year_range}.\n\n"
                        ]
                
                # row.to_excel('Invoice.xlsx', header=heading)  # => proper working, but heading is not displayed in mobile devices, and downloaded file
                
                # row.to_excel('Invoice.xlsx', header=my_heads, startrow=5, startcol=3, index=False)###

                # Exclude the cc_member_name column from the row
                row_without_cc = row.drop(labels=['CC MEMBERS'])
                excel_buffer = row_to_excel(row_without_cc, heading)

                # excel_buffer = row_to_excel(row, heading)###
                
                with open('invoice.xlsx', 'wb') as f:
                    f.write(excel_buffer.getbuffer())

                # message = get_email_template(row, fiscal_year, fiscal_year_range)#for dynamic date

                message = get_email_template(row)

                # Process cc_member_name to create a list of email addresses
                cc_list = [email.strip() for email in row['CC MEMBERS'].split(',')] if pd.notna(row['CC MEMBERS']) else []

                email_background_worker(row['EMAIL'], message, row['PARTY NAME'], cc_list)
                
                # message = get_email_template(row)


            context["message"] = "Email sending in progress"

        except Exception as e:
            import traceback
            traceback.print_exc()
            context["message"] = str(e)

    return render(request, 'pages/index.html', context)



def get_email_template(row):
# def get_email_template(row, fiscal_year_range, fiscal_year): #for dyanamic year

    # today = datetime.today() #For getting the last date current date.
    # last_day_of_month = calendar.monthrange(today.year, today.month)[1] 
    # end_date = today.replace(day=last_day_of_month).strftime('%d %B %Y')

    # {fiscal_year_range} and {fiscal_year}
    

    # Get today's date
    today = datetime.today()

    # Calculate the last day of the previous month
    first_day_of_current_month = datetime(today.year, today.month, 1)
    last_day_of_previous_month = first_day_of_current_month - timedelta(days=1)

    # Format the date
    end_date = last_day_of_previous_month.strftime('%d %B %Y')
    msg = f"""
            <html>
            <body>
            
            <p>Dear Sir/Madam,</p>
        
            <p>During the reconciliation of <b><mark>GSTR-2B for GSTR-2B for FY 2023-2025</mark></b>, we found following discrepancies<br> 
            related to <b>invoices/ Debit Note/ Credit Note</b> not uploaded by you on the GST portal<br> 
            during <b>the April 23 to April 24 F.Y. 23-25.</b> Due to this we are unable to take input tax credit <br>
            of said invoices/ Debit Note/ Credit Notes.  These discrepancies have been worked out on <br>
            the basis of 2B downloaded up to <b>{end_date}</b></p>

            <p>Request to please resolve these issues and submit necessary documents at the earliest for <br>
            said discrepancies as otherwise any reversal of Input tax credit on account of mismatch in <br>
            record will be debited to you along with interest and penalty. This may be due to an error in <br>
            uploading the details of Invoices/ Debit Note/ Credit Note in Name of <b>SHREE GANESH</b> <br>
            <b>PRESS N COAT INDUSTRIES PRIVATE LIMITED</b> on GST No. <b>27AAFCS1275F1ZE.</b></p>
            
            <p>The summary of discrepancies and invoice wise detail is annexed herewith.</p>

            <p>In case of any doubt/ clarification please feel free to write to us.</p>

            <p><b><i>Manesh G Sheolikar</i></b><br>
            <b><i>Shree Ganesh Press N Coat Industries Pvt Ltd</i></b><br>
            <b><i>Account Department</i></b><br>
            <b><i>M-152,Waluj-Aurangabad</i></b><br>
            <b><i>Mob-Number-9881621443</i></b><br>
            <img src="cid:media/save tree.jpg" alt="Save Trees"><br>
            </p>
        </body>
        </html>
            """
    return msg

def row_to_excel(row, heading):
    # Create a DataFrame from the row
    df = pd.DataFrame([row])

    # Create an in-memory buffer
    excel_buffer = BytesIO()

    empty_rows = 3

    # Write the DataFrame to the buffer
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        writer.sheets['Sheet1'] = writer.book.create_sheet('Sheet1')###
        sheet = writer.sheets['Sheet1']

        wb = Workbook()
        ws = wb.active 
         
        for idx, line in enumerate(heading, start=1 + empty_rows):
            sheet.cell(row=idx, column=14, value=line.strip()).alignment = openpyxl.styles.Alignment(horizontal='center')#for making data horrizantal from vertical.
            sheet.cell(row=idx, column=14).font = Font(bold=True)
        
        fill_color = PatternFill(fill_type='solid', fgColor='00FFCC99')#=>Set color to row
        new_color = PatternFill(fill_type='solid', fgColor="FFFF00")
        row_number = 4 + empty_rows #number of row
        
        # max_col = df.shape[0]
        
        # Apply the fill color to the entire row
        # for col in range(1, max_col + 1):
        #     sheet.cell(row=row_number, column=6).fill = fill_color
        
        for col in sheet.iter_rows(min_col=6, max_col=20 + df.shape[0], min_row=row_number, max_row=4 + empty_rows):
            for cell in col:
                cell.fill = fill_color

        for col in sheet.iter_rows(min_col=11, max_col=10 + df.shape[0],min_row=5 + empty_rows, max_row=5 + empty_rows ):
            for cell in col:
                cell.fill = new_color

        for col in sheet.iter_rows(min_col=19, max_col=18 + df.shape[0],min_row=5 + empty_rows, max_row=5 + empty_rows ):
            for cell in col:
                cell.fill = new_color

        # Define a border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply the border to all cells in the DataFrame
        # for row in sheet.iter_rows(min_row=1, max_row=5 + empty_rows + df.shape[0], min_col=6, max_col=21):
        #     for cell in row:
        #         cell.border = thin_border


        # Apply the border to all cells in the DataFrame and heading, skipping empty rows
        start_border_row = 1 + empty_rows
        end_border_row = start_border_row + len(heading) + df.shape[0] + 1
        for row in sheet.iter_rows(min_row=start_border_row, max_row=end_border_row, min_col=6, max_col=21):
            for cell in row:
                cell.border = thin_border


        startrow = len(heading) + 1 + empty_rows
        df.to_excel(writer, index=False, startrow=startrow, startcol=5)

    # Seek to the beginning of the stream
    excel_buffer.seek(0)

    return excel_buffer

def bold_text(text):
  """Applies bold formatting to a string."""
  font = Font(bold=True)
  return font + text

            # This is testing email Please Ignore It !!!
        
            # Hello {row['vendor_name']}, 

            # Your last invoice is pending!
            # The amount is {row['invoice']}.\n\n
            # Regards,
            # Zencon Infotech Pvt Ltd.

